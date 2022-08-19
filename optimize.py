# -*- coding:utf-8 -*-
# 1. 使用预训练模型RoBERTa-wwm-ext-large
# 2. 使用warmup策略
# 3. set param.requires_grad = True
from sklearn.model_selection import train_test_split
from transformers import BertTokenizer
from transformers import BertModel
from transformers import AutoModel
import torch
from torch import nn
import numpy as np
import pandas as pd
from transformers import BertForSequenceClassification, AdamW, BertConfig, get_linear_schedule_with_warmup
from torch.utils.data import TensorDataset, DataLoader, RandomSampler, SequentialSampler
import openpyxl
from sklearn.utils import class_weight
import sys
from sklearn.metrics import classification_report
import random
import transformers
transformers.logging.set_verbosity_error()
ENABLE_GPU = True
PADDING_LENGTH = 512

print('\nDefining hyperparameters\n')
epochs = 30
bs_list = [2, 8, 16, 32]
lr_list = [1e-5, 5e-5, 1e-4]
grad_acc_list = [2,4,8]
batch_size = bs_list[int(sys.argv[1])]
learning_rate = lr_list[int(sys.argv[2])]
grad_accumulate = grad_acc_list[int(sys.argv[3])] 

# model_name = "bert-base-chinese"
model_name = sys.argv[4]
path = 'saved_models/lr='+str(learning_rate)+"bs="+str(batch_size)+"g_a="+str(grad_accumulate)+'.pt' 

if ENABLE_GPU:
    device = torch.device("cuda")

wb = openpyxl.load_workbook('data/data_clean_all.xlsx')
ws = wb.active
serial = []
label = []
for col in ws['A']:
    serial.append(col.value) 
for col in ws['L']:
    label.append(col.value)

sexdict = {1: "男", 2: "女", "性别 \n男=1  女=2": -1, None: -1}
leveldict = {1: "小学", 2: "初中", 3: "高中", 4: "职高", "学段 1=小学 2=初中 3=高中 4=职高": -1, None: -1}
sex,level = [], []
for col in ws["B"]:
    sex.append(sexdict[col.value])
for col in ws["C"]:
    level.append(leveldict[col.value])

scores = []
for col in ws["F"]:
    scores.append(col.value)

data = []
print("  \nReading files to construct raw dataset\n   ")
err4num, errnum = 0, 0
for serial_num, la, se, lev, soft in list(zip(serial, label, sex, level, scores))[1:]:
    if str(serial_num)[0] == "4":
        err4num += 1
        continue
    f = open('data/text_pool/' + str(serial_num) + ".txt", 'r')
    lines = []
    title = f.readline()
    if len(title) > 20:
        lines.append(title)
        title = ""
    for line in f:
        l = line.strip()
        if len(l) > 0:
            lines.append(l)
    if len(lines) == 0:
        errnum += 1
        continue
    onedata = {"title": title, "text": "\n".join(lines), "label": la, "soft_score": soft}
    if len(onedata["text"]) < 15:
        errnum += 1
        continue
    if len(onedata["title"]) > 0 and onedata["title"][0] == "《" and onedata["title"][-1] == "》":
        onedata["title"] = onedata["title"][1:-1]
    onedata["title"] += "\t性别：%s，年级：%s" % (se, lev)
    data.append(onedata)
    f.close()
print("the number with 4:", err4num, "the number of error:", errnum)

print("\nSplit dataset\n")
random.seed(2022)
random.shuffle(data)

train_data = data[:int(len(data) * 0.7)]
valid_data = data[int(len(data) * 0.7) : int(len(data) * 0.85)]
test_data = data[int(len(data) * 0.85) : ]

print("the number of train data", len(train_data))
print("the number of valid data", len(valid_data))
print("the number of test data", len(test_data))

# class_weights = class_weight.compute_class_weight('balanced', classes = np.unique([d["label"] for d in train_data]), y = [d["label"] for d in train_data])


# define bert model
tokenizer = BertTokenizer.from_pretrained(model_name)

def convert_text(titles, texts):
    input_ids, attention_mask = [], []
    for t, txt in zip(titles, texts):
        tokens = [tokenizer.cls_token_id] + tokenizer.encode(t, add_special_tokens=False) + [tokenizer.sep_token_id] + tokenizer.encode(txt, add_special_tokens=False) + [tokenizer.sep_token_id]
        mask = [1] * len(tokens) + [0] * (PADDING_LENGTH - len(tokens))
        tokens += [tokenizer.pad_token_id] * (PADDING_LENGTH - len(tokens))
        input_ids.append(tokens[:PADDING_LENGTH]), attention_mask.append(mask[:PADDING_LENGTH])
    return {"input_ids": input_ids, "attention_mask": attention_mask}

tokens_train = convert_text([d["title"] for d in train_data], [d["text"] for d in train_data])
tokens_val = convert_text([d["title"] for d in valid_data], [d["text"] for d in valid_data])
tokens_test = convert_text([d["title"] for d in test_data], [d["text"] for d in test_data])
'''
# tokenize and encode sequences in the training set
tokens_train = tokenizer(
    [d["title"] for d in train_data],
    text_pair = [d["text"] for d in train_data],
    max_length = PADDING_LENGTH,
    padding='max_length',
    truncation=True
)

# tokenize and encode sequences in the validation set
tokens_val = tokenizer(
    [d["title"] for d in valid_data],
    text_pair = [d["text"] for d in valid_data],
    max_length = PADDING_LENGTH,
    padding='max_length',
    truncation=True
)

# tokenize and encode sequences in the test set
tokens_test = tokenizer(
    [d["title"] for d in test_data],
    text_pair = [d["text"] for d in test_data],
    max_length = PADDING_LENGTH,
    padding='max_length',
    truncation=True
)
'''

print('\nconvert lists to tensors\n')

train_seq = torch.tensor(tokens_train['input_ids'], dtype=torch.long)
train_mask = torch.tensor(tokens_train['attention_mask'], dtype=torch.long)
train_y = torch.tensor([d["label"] if d["soft_score"] >= 53 or d["soft_score"] <= 40 else -100 for d in train_data ], dtype=torch.long)

val_seq = torch.tensor(tokens_val['input_ids'], dtype=torch.long)
val_mask = torch.tensor(tokens_val['attention_mask'], dtype=torch.long)
val_y = torch.tensor([d["label"] for d in valid_data], dtype=torch.long)

test_seq = torch.tensor(tokens_test['input_ids'], dtype=torch.long)
test_mask = torch.tensor(tokens_test['attention_mask'], dtype=torch.long)
test_y = torch.tensor([d["label"] for d in test_data], dtype=torch.long)

train_data = TensorDataset(train_seq, train_mask, train_y)
train_sampler = RandomSampler(train_data)
train_dataloader = DataLoader(train_data, sampler=train_sampler, batch_size=batch_size)

val_data = TensorDataset(val_seq, val_mask, val_y)
val_sampler = SequentialSampler(val_data)
val_dataloader = DataLoader(val_data, sampler = val_sampler, batch_size=batch_size)

test_data = TensorDataset(test_seq, test_mask, test_y)
test_sampler = SequentialSampler(test_data)
test_dataloader = DataLoader(test_data, sampler = test_sampler, batch_size=batch_size, drop_last=False)


# # unfreeze
# for param in bert.parameters():
#     param.requires_grad = True

# model definition
class BERT_Arch(nn.Module):
    def __init__(self):
        super(BERT_Arch, self).__init__()
        self.bert = BertModel.from_pretrained(model_name)
        # self.dropout = nn.Dropout(0.3)
        # self.relu =  nn.ReLU()
        # self.fc1 = nn.Linear(1024,512)
        self.fc2 = nn.Linear(self.bert.config.hidden_size, 2) # binary classification

    #define the forward pass
    def forward(self, sent_id, mask):
        out = self.bert(sent_id, attention_mask=mask)
        x = self.fc2(out["pooler_output"])
        return x

print("\nLoad model\n")
model = BERT_Arch()
# model.load_state_dict(torch.load('saved_models/lr=5e-05bs=8g_a=4.pt'))
if ENABLE_GPU:
    model = model.to(device)

optimizer = AdamW(model.parameters(), lr = learning_rate)
cross_entropy  = nn.CrossEntropyLoss(ignore_index=-100)#weight=torch.FloatTensor(class_weights).to(device))

print("len(train_data)",len(train_data))
total_step = (len(train_dataloader)) * epochs
warm_up_ratio = 0

# scheduler = get_linear_schedule_with_warmup(optimizer,
#                                             num_warmup_steps= 0,
#                                             num_training_steps=total_step)

def train():
    print('...Training model...')
    model.train()
    total_loss = 0
    # empty list to save model predictions
    total_preds=[]
    # iterate over batches
    # correct = 0
    for step,batch in enumerate(train_dataloader):
        # push the batch to gpu
        if ENABLE_GPU:
            batch = [r.to(device) for r in batch]
        sent_id, mask, labels = batch      
        preds = model(sent_id, mask)
        loss = cross_entropy(preds, labels) # compute loss
        total_loss = total_loss + loss.item()
        loss = loss / grad_accumulate 
        loss.backward() # backward propagation
        # torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)
        if ((step + 1) % grad_accumulate == 0) or (step == (len(train_dataloader) - 1)):
            optimizer.step() # update model
            optimizer.zero_grad() # clear gradient to zero
    # compute the training loss of the epoch
    avg_loss = total_loss / len(train_dataloader)
    # total_preds  = np.concatenate(total_preds, axis=0)
    return avg_loss, total_preds

# function for evaluating the model
def evaluate(dataloader):
    print("Evaluating...")
    # deactivate dropout layers
    model.eval() # disable the dropout layer for evaluating
    right, total = 0, 0
    # empty list to save the model predictions
    total_preds = []
    for step,batch in enumerate(dataloader):
        # if step % 50 == 0 and not step == 0:
        #     print('Batch {:>5,}  of  {:>5,}.'.format(step, len(dataloader)))
        if ENABLE_GPU:
            batch = [t.to(device) for t in batch]
        sent_id, mask, labels = batch
        with torch.no_grad():
            scores = model(sent_id, mask) # make predictions
            pred = torch.max(scores, dim=1)[1]

            right += (pred[labels >= 0] == labels[labels >= 0]).sum().item()
            total += int(labels[labels >= 0].size(0))

    avg_acc = right / total
    # total_preds  = np.concatenate(total_preds, axis=0)
    return avg_acc, total_preds

def fine_tunning():
    # set initial acc to zero
    best_valid_acc = 0
    # empty lists to store training and validation loss of each epoch
    train_losses=[]
    valid_acc_list =[] # store validation accuracy
    saving_cnt = 0 # count storing model times 
    #for each epoch
    for epoch in range(epochs):
        
        print('========== Epoch {:} / {:} =========='.format(epoch + 1, epochs))
        #train model
        train_loss, _ = train()
        
        #evaluate model
        valid_acc, _ = evaluate(val_dataloader)
        test_acc, _ = evaluate(test_dataloader)
        print("Training Loss:", train_loss, "valid_acc:", valid_acc, "test_acc:", test_acc)
        #save the best model
        if valid_acc > best_valid_acc:
            best_valid_acc = valid_acc
            print("saving the {}.th model, with lr = {}, bs = {}, grad_accumulate ={}\n".format(saving_cnt, learning_rate, batch_size, grad_accumulate))
            torch.save(model.state_dict(), path) # save the newest model
            saving_cnt += 1
        
        train_losses.append(train_loss)
        valid_acc_list.append(valid_acc)
    # 将valid_acc_list保存下来,用于日后画图使用
    df = pd.DataFrame(valid_acc_list)
    csv_name = 'saved_models/lr='+str(learning_rate)+"bs="+str(batch_size)+"g_a="+str(grad_accumulate)+".csv" 
    df.to_csv(csv_name)


if __name__ == "__main__":
    fine_tunning()

